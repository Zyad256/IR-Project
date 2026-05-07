"""
Indexer Module - Handles file parsing and Elasticsearch indexing.

File Format Handling:
- TXT: Full file content as one document.
- PDF: Full extracted text as one document.
- JSON: If the file contains a list of objects, each object becomes a separate document
        with all string values concatenated. If it's a single object, it becomes one document.
- CSV: Each row becomes a separate document. All column values are concatenated as the text.
- Excel (.xlsx): Each row becomes a separate document. All cell values are concatenated as the text.
"""

import os
import json
import csv
import hashlib
from datetime import datetime
from typing import List, Dict, Optional

from elasticsearch import Elasticsearch, helpers
from PyPDF2 import PdfReader
from openpyxl import load_workbook


INDEX_NAME = "search_engine_docs"

# Elasticsearch mapping with proper analyzers for suggestions
INDEX_SETTINGS = {
    "settings": {
        "number_of_shards": 1,
        "number_of_replicas": 0,
        "analysis": {
            "analyzer": {
                "custom_analyzer": {
                    "type": "custom",
                    "tokenizer": "standard",
                    "filter": ["lowercase", "stop"]
                }
            }
        }
    },
    "mappings": {
        "properties": {
            "filename": {"type": "keyword"},
            "filepath": {"type": "keyword"},
            "file_type": {"type": "keyword"},
            "content": {
                "type": "text",
                "analyzer": "custom_analyzer",
                "term_vector": "with_positions_offsets",
                "fielddata": True
            },
            "modification_date": {"type": "date"},
            "file_size": {"type": "long"},
            "sub_doc_id": {"type": "keyword"},  # for CSV/JSON rows
        }
    }
}


def get_es_client() -> Elasticsearch:
    """Create and return an Elasticsearch client."""
    es = Elasticsearch(
        "http://localhost:9200",
        request_timeout=120
    )
    return es


def create_index(es: Elasticsearch, force_rebuild: bool = False):
    """Create the Elasticsearch index with proper mappings."""
    if es.indices.exists(index=INDEX_NAME):
        if force_rebuild:
            es.indices.delete(index=INDEX_NAME)
        else:
            return
    es.indices.create(index=INDEX_NAME, body=INDEX_SETTINGS)


def parse_txt(filepath: str) -> List[Dict]:
    """Parse a TXT file - entire content as one document."""
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    return [{"content": content}]


def parse_pdf(filepath: str) -> List[Dict]:
    """Parse a PDF file - all pages concatenated as one document."""
    try:
        reader = PdfReader(filepath)
        text_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
        content = "\n".join(text_parts)
        return [{"content": content}] if content.strip() else []
    except Exception as e:
        print(f"Error parsing PDF {filepath}: {e}")
        return [{"content": f"[PDF parse error: {e}]"}]


def parse_json(filepath: str) -> List[Dict]:
    """
    Parse a JSON file.
    - If it's a list of objects: each object = one document, all string values concatenated.
    - If it's a single object: one document with all string values concatenated.
    """
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        data = json.load(f)

    def extract_text(obj, prefix=""):
        """Recursively extract all string values from a JSON object."""
        parts = []
        if isinstance(obj, dict):
            for key, value in obj.items():
                parts.extend(extract_text(value, f"{prefix}{key}: "))
        elif isinstance(obj, list):
            for item in obj:
                parts.extend(extract_text(item, prefix))
        elif isinstance(obj, str):
            parts.append(f"{prefix}{obj}")
        elif obj is not None:
            parts.append(f"{prefix}{str(obj)}")
        return parts

    documents = []
    if isinstance(data, list):
        for i, item in enumerate(data):
            text = " | ".join(extract_text(item))
            if text.strip():
                documents.append({"content": text, "sub_doc_id": f"row_{i}"})
    else:
        text = " | ".join(extract_text(data))
        if text.strip():
            documents.append({"content": text})

    return documents


def parse_csv(filepath: str) -> List[Dict]:
    """Parse a CSV file - each row becomes a separate document."""
    documents = []
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for i, row in enumerate(reader):
            parts = []
            for header in headers:
                value = row.get(header, "")
                if value and str(value).strip():
                    parts.append(f"{header}: {value}")
            text = " | ".join(parts)
            if text.strip():
                documents.append({"content": text, "sub_doc_id": f"row_{i}"})
    return documents


def parse_excel(filepath: str) -> List[Dict]:
    """Parse an Excel (.xlsx) file - each row becomes a separate document."""
    documents = []
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h else f"col_{j}" for j, h in enumerate(rows[0])]
            for i, row in enumerate(rows[1:], start=1):
                parts = []
                for j, cell in enumerate(row):
                    if cell is not None and str(cell).strip():
                        header = headers[j] if j < len(headers) else f"col_{j}"
                        parts.append(f"{header}: {cell}")
                text = " | ".join(parts)
                if text.strip():
                    documents.append({
                        "content": text,
                        "sub_doc_id": f"{sheet_name}_row_{i}"
                    })
        wb.close()
    except Exception as e:
        print(f"Error parsing Excel {filepath}: {e}")
        return [{"content": f"[Excel parse error: {e}]"}]
    return documents


# Map file extensions to parsers
PARSERS = {
    ".txt": parse_txt,
    ".pdf": parse_pdf,
    ".json": parse_json,
    ".csv": parse_csv,
    ".xlsx": parse_excel,
}


def get_file_extension(filepath: str) -> str:
    """Get the lowercase file extension."""
    return os.path.splitext(filepath)[1].lower()


def scan_folder(folder_path: str, allowed_extensions: List[str]) -> List[str]:
    """Scan a folder recursively for files with allowed extensions."""
    files = []
    for root, dirs, filenames in os.walk(folder_path):
        for filename in filenames:
            ext = get_file_extension(filename)
            if ext in allowed_extensions:
                files.append(os.path.join(root, filename))
    return files


def index_files(
    es: Elasticsearch,
    folder_path: str,
    allowed_formats: List[str],
    force_rebuild: bool = True
) -> Dict:
    """
    Index files from a folder into Elasticsearch.
    
    Returns a summary dict with counts by file type.
    """
    create_index(es, force_rebuild=force_rebuild)

    # Normalize extensions
    allowed_extensions = [f".{fmt.lower().strip('.')}" for fmt in allowed_formats]

    # Scan for files
    files = scan_folder(folder_path, allowed_extensions)

    stats = {
        "total_files": 0,
        "total_documents": 0,
        "by_type": {},
        "errors": []
    }

    actions = []

    for filepath in files:
        ext = get_file_extension(filepath)
        parser = PARSERS.get(ext)
        if not parser:
            stats["errors"].append(f"No parser for {ext}: {filepath}")
            continue

        try:
            documents = parser(filepath)
        except Exception as e:
            stats["errors"].append(f"Error parsing {filepath}: {e}")
            continue

        file_stat = os.stat(filepath)
        mod_time = datetime.fromtimestamp(file_stat.st_mtime).isoformat()
        file_size = file_stat.st_size
        filename = os.path.basename(filepath)
        file_type = ext.lstrip(".")

        stats["total_files"] += 1
        stats["by_type"][file_type] = stats["by_type"].get(file_type, 0) + 1

        for doc in documents:
            # Create a unique ID for each document
            doc_id_source = f"{filepath}_{doc.get('sub_doc_id', 'main')}"
            doc_id = hashlib.md5(doc_id_source.encode()).hexdigest()

            action = {
                "_index": INDEX_NAME,
                "_id": doc_id,
                "_source": {
                    "filename": filename,
                    "filepath": filepath,
                    "file_type": file_type,
                    "content": doc["content"],
                    "modification_date": mod_time,
                    "file_size": file_size,
                    "sub_doc_id": doc.get("sub_doc_id", "main"),
                }
            }
            actions.append(action)
            stats["total_documents"] += 1

    # Bulk index
    if actions:
        success, errors = helpers.bulk(es, actions, raise_on_error=False)
        if errors:
            stats["errors"].extend([str(e) for e in errors[:10]])

    # Refresh index to make documents searchable immediately
    es.indices.refresh(index=INDEX_NAME)

    return stats
